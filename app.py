import os
from datetime import timedelta, datetime
from dotenv import load_dotenv
from supabase import create_client, Client
from functools import wraps
import secrets
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, make_response, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from flask_wtf.csrf import CSRFProtect
from html import escape
import logging
import re
import requests
from uuid import uuid4
import pandas as pd
from io import BytesIO
import pytz
from collections import Counter
from openpyxl.chart.label import DataLabelList
import json

# Load environment variables
load_dotenv()

app = Flask(__name__)
app.config['DEBUG'] = True
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', secrets.token_hex(24))
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)

# Initialize database
db = SQLAlchemy(app)
migrate = Migrate(app, db)

# Initialize CSRF protection
csrf = CSRFProtect(app)
csrf.exempt("api_login")
csrf.exempt("get_user_info")

# Configure logging
logging.basicConfig(filename='app.log', level=logging.INFO, format='%(asctime)s %(levelname)s: %(message)s')
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

# Supabase setup
SUPABASE_URL = os.environ.get('SUPABASE_URL')
SUPABASE_KEY = os.environ.get('SUPABASE_KEY')
SUPABASE_SERVICE_KEY = os.environ.get('SUPABASE_SERVICE_KEY')

if not all([SUPABASE_URL, SUPABASE_KEY, SUPABASE_SERVICE_KEY]):
    raise ValueError("SUPABASE_URL, SUPABASE_KEY, and SUPABASE_SERVICE_KEY must be set")

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
supabase_service: Client = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)

# Valid domains
VALID_CATEGORIES = ['Longhaul', 'GPON_FMC']
VALID_DOMAINS = {
    'Longhaul': [
        'FMC Taxila', 'FMC Fateh Jang', 'FMC Rawalpindi', 'FMC Murree',
        'FMC Gujar Khan', 'FMC Chakwal', 'FMC Talagang', 'FMC Jhelum', 'FMC PD Khan'
    ],
    'GPON_FMC': [
        'FMC Attock GPON', 'FMC Wah GPON', 'FMC Taxila GPON', 'FMC Murree GPON',
        'FMC Gujar Khan GPON', 'FMC Chakwal GPON', 'FMC Jhelum GPON'
    ]
}

# Helper functions
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        auth_token = request.cookies.get('auth_token')
        if auth_token:
            try:
                response = supabase.auth.get_user(auth_token)
                if response.user:
                    user_data = supabase_service.table('users_info').select('user_id', 'username', 'region', 'category', 'domain', 'role').eq('user_id', response.user.id).execute()
                    if user_data.data:
                        session['user_id'] = user_data.data[0]['user_id']
                        session['username'] = user_data.data[0]['username']
                        session['region'] = user_data.data[0]['region']
                        session['category'] = user_data.data[0]['category']  # Now a list
                        session['domain'] = user_data.data[0]['domain']      # Now a list
                        session['role'] = user_data.data[0]['role']
                        return f(*args, **kwargs)
            except Exception as e:
                logger.error(f"Token validation failed: {str(e)}")
                resp = make_response(redirect(url_for('login')))
                resp.set_cookie('auth_token', '', expires=0)
                session.clear()
                flash('Invalid or expired session. Please log in again.')
                return resp
        else:
            flash('Please log in to access this page')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def master_required(f):
    @wraps(f)
    @login_required
    def decorated_function(*args, **kwargs):
        if session.get('role') != 'master':
            flash('Access restricted to master users')
            return redirect(url_for('index'))
        # For master users, ensure they have access to all categories and domains
        all_categories = set(VALID_CATEGORIES)
        all_domains = set(sum(VALID_DOMAINS.values(), []))
        user_categories = set(session.get('category', []))
        user_domains = set(session.get('domain', []))
        if not (all_categories.issubset(user_categories) and all_domains.issubset(user_domains)):
            flash('Master users must have access to all categories and domains')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def safe_float(value, field_name):
    try:
        return float(value) if value.strip() else None
    except ValueError:
        raise ValueError(f"Invalid numeric value for {field_name}: {value}")

def safe_int(value, field_name):
    try:
        return int(value) if value.strip() else None
    except ValueError:
        raise ValueError(f"Invalid integer value for {field_name}: {value}")

def sanitize_text(value):
    return escape(value.strip()) if value else None

def validate_email(email):
    email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(email_regex, email) is not None

def apply_data_filter(query):
    user_role = session.get('role')
    user_categories = session.get('category', [])
    user_domains = session.get('domain', [])

    logger.debug(f"Applying filter: role={user_role}, categories={user_categories}, domains={user_domains}")

    if not all([user_role, user_categories, user_domains]):
        logger.error("Missing session data for filtering")
        raise ValueError("User session data incomplete")

    if user_role == 'master':
        if set(VALID_CATEGORIES).issubset(set(user_categories)) and set(sum(VALID_DOMAINS.values(), [])).issubset(set(user_domains)):
            logger.debug("No filter applied (master with all categories and domains)")
            return query
        else:
            # Filter by user's assigned categories and domains
            valid_domains = []
            for cat in user_categories:
                if cat in VALID_CATEGORIES:
                    valid_domains.extend(VALID_DOMAINS.get(cat, []))
            valid_domains = [d for d in valid_domains if d in user_domains]
            if not valid_domains:
                logger.error(f"No valid domains for categories: {user_categories}")
                raise ValueError("Invalid categories or domains")
            logger.debug(f"Filtering by categories={user_categories}, domains={valid_domains}")
            return query.filter(
                FMCInformation.category.in_(user_categories),
                FMCInformation.domain.in_(valid_domains)
            )
    else:
        # Non-master users: filter by their assigned categories and domains
        valid_domains = []
        for cat in user_categories:
            if cat in VALID_CATEGORIES:
                valid_domains.extend(VALID_DOMAINS.get(cat, []))
        valid_domains = [d for d in valid_domains if d in user_domains]
        if not valid_domains:
            logger.error(f"No valid domains for user categories: {user_categories}")
            raise ValueError("Invalid categories or domains for user")
        logger.debug(f"Non-master: Filtering by categories={user_categories}, domains={valid_domains}")
        return query.filter(
            FMCInformation.category.in_(user_categories),
            FMCInformation.domain.in_(valid_domains)
        )
    
# Database Models
class FMCInformation(db.Model):
    __tablename__ = 'fmc_information'
    id = db.Column(db.Integer, primary_key=True)
    region = db.Column(db.String(50), nullable=False, default='RTR')
    category = db.Column(db.String(50), nullable=False)
    domain = db.Column(db.String(100), nullable=False)
    cable_cut_noc_id = db.Column(db.String(100))
    cable_used_meters = db.Column(db.Float)
    cable_type = db.Column(db.String(50))
    cable_capacity = db.Column(db.String(50))
    no_of_joints = db.Column(db.Integer)
    created_by = db.Column(db.String(50), nullable=False)
    updated_by = db.Column(db.String(50), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow)
    joint_types = db.relationship('JointType', backref='fmc_info', lazy=True, cascade="all, delete-orphan")
    pipe_info = db.relationship('PipeInformation', backref='fmc_info', lazy=True, cascade="all, delete-orphan")

class JointType(db.Model):
    __tablename__ = 'joint_types'
    id = db.Column(db.Integer, primary_key=True)
    fmc_id = db.Column(db.Integer, db.ForeignKey('fmc_information.id'), nullable=False)
    joint_type = db.Column(db.String(100))
    created_by = db.Column(db.String(50), nullable=False)
    updated_by = db.Column(db.String(50), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow)

class PipeInformation(db.Model):
    __tablename__ = 'pipe_information'
    id = db.Column(db.Integer, primary_key=True)
    fmc_id = db.Column(db.Integer, db.ForeignKey('fmc_information.id'), nullable=False)
    pipe_used_meters = db.Column(db.Float)
    pipe_size_inches = db.Column(db.Float)
    pipe_type = db.Column(db.String(100))
    created_by = db.Column(db.String(50), nullable=False)
    updated_by = db.Column(db.String(50), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow)

# Routes
@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        username = request.form.get('username', '').lower().strip()
        email = request.form.get('email', '').lower().strip()
        password = request.form.get('password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        categories = request.form.getlist('category[]')  # Multiple categories
        domains = request.form.getlist('domain[]')       # Multiple domains
        role = request.form.get('role', 'user').strip()

        if not all([username, email, password, confirm_password, categories, domains]):
            flash('All fields are required')
            return render_template('signup.html')

        if not validate_email(email):
            flash('Invalid email format')
            return render_template('signup.html')

        if password != confirm_password:
            flash('Passwords do not match')
            return render_template('signup.html')

        if role not in ['user', 'master']:
            flash('Invalid role')
            return render_template('signup.html')

        # Validate categories and domains
        for cat in categories:
            if cat not in VALID_CATEGORIES:
                flash(f'Invalid category: {cat}')
                return render_template('signup.html')
        valid_domains = []
        for cat in categories:
            valid_domains.extend(VALID_DOMAINS.get(cat, []))
        for dom in domains:
            if dom not in valid_domains:
                flash(f'Invalid domain for selected categories: {dom}')
                return render_template('signup.html')

        if role == 'master':
            all_categories = set(VALID_CATEGORIES)
            all_domains = set(sum(VALID_DOMAINS.values(), []))
            if not (set(categories) == all_categories and set(domains) == all_domains):
                flash('Master users must have access to all categories and domains')
                return render_template('signup.html')

        try:
            user_data = supabase_service.table('users_info').select('username').eq('username', username).execute()
            if user_data.data:
                flash('Username already exists')
                return render_template('signup.html')

            try:
                existing = supabase.auth.admin.get_user_by_email(email)
                if existing.user:
                    flash('Email already registered')
                    return render_template('signup.html')
            except Exception:
                pass

            response = supabase.auth.sign_up({
                "email": email,
                "password": password,
                "options": {
                    "data": {
                        "username": username,
                        "category": categories,  # Store as array
                        "domain": domains,       # Store as array
                        "region": "RTR",
                        "role": role
                    }
                }
            })

            if response.user:
                supabase_service.table('users_info').insert({
                    'user_id': response.user.id,
                    'username': username,
                    'email': email,
                    'region': 'RTR',
                    'category': categories,
                    'domain': domains,
                    'role': role
                }).execute()
                flash('Signup successful! Please check your email to verify your account.')
                return redirect(url_for('login'))
            else:
                flash('Signup failed')
                return render_template('signup.html')
        except Exception as e:
            flash(f'Signup failed: {str(e)}')
            return render_template('signup.html')

    return render_template('signup.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').lower().strip()
        password = request.form.get('password', '').strip()

        if not username or not password:
            flash('Username and password are required')
            return render_template('login.html')

        try:
            user_data = supabase_service.table('users_info').select('user_id', 'region', 'category', 'domain', 'role').eq('username', username).execute()
            if not user_data.data:
                flash('Invalid username')
                return render_template('login.html')

            user_id = user_data.data[0]['user_id']
            region = user_data.data[0]['region']
            categories = user_data.data[0]['category']  # Now a list
            domains = user_data.data[0]['domain']      # Now a list
            role = user_data.data[0]['role']

            auth_user = supabase_service.auth.admin.get_user_by_id(user_id)
            if not auth_user.user:
                flash('User not found in authentication system')
                return render_template('login.html')

            email = auth_user.user.email
            email_confirmed = bool(auth_user.user.email_confirmed_at)

            if not email_confirmed:
                supabase.auth.resend({"type": "signup", "email": email})
                flash('Please verify your email. A new verification email has been sent.')
                return render_template('login.html')

            response = supabase.auth.sign_in_with_password({"email": email, "password": password})
            if response.user and response.user.id == user_id:
                session['region'] = region
                session['username'] = username
                session['user_id'] = user_id
                session['category'] = categories
                session['domain'] = domains
                session['role'] = role
                resp = make_response(redirect(url_for('index')))
                resp.set_cookie(
                    'auth_token',
                    response.session.access_token,
                    httponly=True,
                    secure=True,
                    samesite='Lax',
                    max_age=int(timedelta(hours=24).total_seconds())
                )
                return resp
            else:
                flash('Invalid credentials')
                return render_template('login.html')
        except Exception as e:
            flash(f'Login failed: {str(e)}')
            return render_template('login.html')

    return render_template('login.html')

@app.route('/password_reset', methods=['GET', 'POST'])
def password_reset():
    if request.method == 'POST':
        username = request.form.get('username', '').lower().strip()

        if not username:
            flash('Username is required')
            return render_template('password_reset.html')

        try:
            user_data = supabase_service.table('users_info').select('user_id').eq('username', username).execute()
            if not user_data.data:
                flash('Username does not exist')
                return render_template('password_reset.html')

            user_id = user_data.data[0]['user_id']
            auth_user = supabase_service.auth.admin.get_user_by_id(user_id)
            if not auth_user.user:
                flash('User not found in authentication system')
                return render_template('password_reset.html')

            email = auth_user.user.email
            supabase.auth.reset_password_email(email)
            flash('Password reset email sent! Please check your email inbox.')
            return redirect(url_for('login'))

        except Exception as e:
            flash(f'Password reset failed: {str(e)}')
            return render_template('password_reset.html')

    return render_template('password_reset.html')

@app.route('/password_reset_confirm', methods=['GET', 'POST'])
def password_reset_confirm():
    access_token = request.args.get('access_token') or request.form.get('access_token')
    if request.method == 'POST':
        new_password = request.form.get('new_password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()

        if not new_password or not confirm_password:
            flash('All fields are required')
            return render_template('password_reset_confirm.html', access_token=access_token)
        if new_password != confirm_password:
            flash('Passwords do not match')
            return render_template('password_reset_confirm.html', access_token=access_token)

        if not access_token:
            flash('Invalid or missing access token')
            return redirect(url_for('login'))

        try:
            headers = {
                "apikey": os.environ["SUPABASE_SERVICE_KEY"],
                "Authorization": f"Bearer {access_token}",
                "Content-Type": "application/json",
            }
            data = {
                "password": new_password
            }
            response = requests.put(
                f"{os.environ['SUPABASE_URL']}/auth/v1/user",
                headers=headers,
                json=data
            )
            if response.status_code == 200:
                flash('Password reset successfully! Please log in.')
                return redirect(url_for('login'))
            else:
                error_msg = response.json().get("msg") or response.text
                flash(f'Password reset failed: {error_msg}')
                return render_template('password_reset_confirm.html', access_token=access_token)
        except Exception as e:
            flash(f'Password reset failed: {str(e)}')
            return render_template('password_reset_confirm.html', access_token=access_token)

    if not access_token:
        flash('Invalid or missing access token')
        return redirect(url_for('login'))

    return render_template('password_reset_confirm.html', access_token=access_token)

@app.route('/logout')
@login_required
def logout():
    username = session.get('username')
    resp = make_response(redirect(url_for('login')))
    resp.set_cookie('auth_token', '', expires=0)
    session.clear()
    try:
        supabase.auth.sign_out()
    except Exception as e:
        logger.warning(f"Error during logout for {username}: {str(e)}")
    return resp

@app.route('/')
@login_required
def index():
    try:
        user_role = session.get('role')
        user_categories = session.get('category', [])
        user_domains = session.get('domain', [])
        username = session.get('username')

        logger.debug(f"User: {username}, Role: {user_role}, Categories: {user_categories}, Domains: {user_domains}")

        # Base query for FMC entries
        base_query = FMCInformation.query
        filtered_query = apply_data_filter(base_query)
        entries = filtered_query.order_by(FMCInformation.created_at.desc()).limit(5).all()

        # --- NOC ID BY CATEGORY OVERALL ---
        noc_by_category = (
            filtered_query
            .with_entities(FMCInformation.category, db.func.count(FMCInformation.cable_cut_noc_id))
            .group_by(FMCInformation.category)
            .all()
        )
        noc_by_category_dict = {cat: count for cat, count in noc_by_category}

        # --- NOC ID BY CATEGORY YEAR WISE ---
        noc_by_category_year = (
            filtered_query
            .with_entities(
                db.extract('year', FMCInformation.created_at).label('year'),
                FMCInformation.category,
                db.func.count(FMCInformation.cable_cut_noc_id)
            )
            .group_by('year', FMCInformation.category)
            .order_by('year')
            .all()
        )
        noc_category_year_labels = sorted(set(str(int(x[0])) for x in noc_by_category_year))
        noc_category_year_cats = sorted(set(x[1] for x in noc_by_category_year))
        noc_category_year_data = {
            cat: [0] * len(noc_category_year_labels) for cat in noc_category_year_cats
        }
        for year, cat, count in noc_by_category_year:
            idx = noc_category_year_labels.index(str(int(year)))
            noc_category_year_data[cat][idx] = count

        # --- NOC ID BY CATEGORY MONTH WISE (2025 only) ---
        noc_by_category_month = (
            filtered_query.filter(db.extract('year', FMCInformation.created_at) == 2025)
            .with_entities(
                db.func.to_char(FMCInformation.created_at, 'MM').label('month'),
                FMCInformation.category,
                db.func.count(FMCInformation.cable_cut_noc_id)
            )
            .group_by('month', FMCInformation.category)
            .order_by('month')
            .all()
        )
        month_labels = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        noc_category_month_cats = sorted(set(x[1] for x in noc_by_category_month))
        noc_category_month_data = {cat: [0] * 12 for cat in noc_category_month_cats}
        for month, cat, count in noc_by_category_month:
            noc_category_month_data[cat][int(month)-1] = count

        # --- NOC BY DOMAIN OVERALL ---
        noc_by_domain = (
            filtered_query
            .with_entities(FMCInformation.domain, db.func.count(FMCInformation.cable_cut_noc_id))
            .group_by(FMCInformation.domain)
            .all()
        )
        noc_by_domain_dict = {dom: count for dom, count in noc_by_domain}

        # --- NOC BY DOMAIN YEAR WISE ---
        noc_by_domain_year = (
            filtered_query
            .with_entities(
                db.extract('year', FMCInformation.created_at).label('year'),
                FMCInformation.domain,
                db.func.count(FMCInformation.cable_cut_noc_id)
            )
            .group_by('year', FMCInformation.domain)
            .order_by('year')
            .all()
        )
        noc_domain_year_labels = sorted(set(str(int(x[0])) for x in noc_by_domain_year))
        noc_domain_year_doms = sorted(set(x[1] for x in noc_by_domain_year))
        noc_domain_year_data = {
            dom: [0] * len(noc_domain_year_labels) for dom in noc_domain_year_doms
        }
        for year, dom, count in noc_by_domain_year:
            idx = noc_domain_year_labels.index(str(int(year)))
            noc_domain_year_data[dom][idx] = count

        # --- NOC BY DOMAIN MONTH WISE (2025 only) ---
        noc_by_domain_month = (
            filtered_query.filter(db.extract('year', FMCInformation.created_at) == 2025)
            .with_entities(
                db.func.to_char(FMCInformation.created_at, 'MM').label('month'),
                FMCInformation.domain,
                db.func.count(FMCInformation.cable_cut_noc_id)
            )
            .group_by('month', FMCInformation.domain)
            .order_by('month')
            .all()
        )
        noc_domain_month_doms = sorted(set(x[1] for x in noc_by_domain_month))
        noc_domain_month_data = {dom: [0] * 12 for dom in noc_domain_month_doms}
        for month, dom, count in noc_by_domain_month:
            noc_domain_month_data[dom][int(month)-1] = count

        # --- Existing stats for summary, cable, joints, pipe, etc ---
        total_entries = filtered_query.count()
        longhaul_count = filtered_query.filter(FMCInformation.category == 'Longhaul').count()
        gpon_fmc_count = filtered_query.filter(FMCInformation.category == 'GPON_FMC').count()
        total_cable_used = filtered_query.with_entities(db.func.coalesce(db.func.sum(FMCInformation.cable_used_meters), 0)).scalar()
        total_joints = filtered_query.with_entities(db.func.coalesce(db.func.sum(FMCInformation.no_of_joints), 0)).scalar()
        pipe_query = apply_data_filter(db.session.query(PipeInformation).join(FMCInformation))
        total_pipe_used = pipe_query.with_entities(db.func.coalesce(db.func.sum(PipeInformation.pipe_used_meters), 0)).scalar()

        # Cable type counts
        cable_type_counts = dict(
            filtered_query.filter(FMCInformation.cable_type.isnot(None))
            .with_entities(FMCInformation.cable_type, db.func.count(FMCInformation.cable_cut_noc_id))
            .group_by(FMCInformation.cable_type)
            .all()
        )
        # Pipe size counts
        pipe_size_counts = dict(
            pipe_query.filter(PipeInformation.pipe_size_inches.isnot(None))
            .with_entities(PipeInformation.pipe_size_inches, db.func.count(PipeInformation.id))
            .group_by(PipeInformation.pipe_size_inches)
            .all()
        )

        # Month-wise NOC ID counts for 2025
        month_counts = (
            filtered_query.filter(db.extract('year', FMCInformation.created_at) == 2025)
            .with_entities(
                db.func.to_char(FMCInformation.created_at, 'MM').label('month'),
                db.func.count(FMCInformation.cable_cut_noc_id)
            )
            .group_by('month')
            .order_by('month')
            .all()
        )
        month_data = [0] * 12
        for month, count in month_counts:
            month_data[int(month) - 1] = count

        # Yearly NOC ID totals
        year_counts = (
            filtered_query
            .with_entities(
                db.extract('year', FMCInformation.created_at).label('year'),
                db.func.count(FMCInformation.cable_cut_noc_id)
            )
            .group_by('year')
            .all()
        )
        year_labels = ['2024', '2025']
        year_data = [0] * len(year_labels)
        for year, count in year_counts:
            if str(int(year)) in year_labels:
                year_data[year_labels.index(str(int(year)))] = count

        # Cable capacity distribution
        cable_capacity_rows = (
            filtered_query.filter(FMCInformation.cable_capacity.isnot(None))
            .with_entities(FMCInformation.cable_capacity, db.func.coalesce(db.func.sum(FMCInformation.cable_used_meters), 0))
            .group_by(FMCInformation.cable_capacity)
            .all()
        )
        cable_capacity_labels = [row[0] for row in cable_capacity_rows]
        cable_capacity_data = [float(row[1]) for row in cable_capacity_rows]

        # Month-wise Cable Type counts for 2025
        cable_type_month_counts = (
            filtered_query.filter(
                db.extract('year', FMCInformation.created_at) == 2025,
                FMCInformation.cable_type.isnot(None)
            )
            .with_entities(
                db.func.to_char(FMCInformation.created_at, 'MM').label('month'),
                FMCInformation.cable_type,
                db.func.count(FMCInformation.cable_cut_noc_id)
            )
            .group_by('month', FMCInformation.cable_type)
            .order_by('month')
            .all()
        )
        cable_types = sorted(set(row[1] for row in cable_type_month_counts))
        cable_type_month_data = {c_type: [0] * 12 for c_type in cable_types}
        for month, c_type, count in cable_type_month_counts:
            cable_type_month_data[c_type][int(month) - 1] = count

        # Month-wise Pipe Size counts for 2025
        pipe_size_month_counts = (
            pipe_query.filter(
                db.extract('year', FMCInformation.created_at) == 2025,
                PipeInformation.pipe_size_inches.isnot(None)
            )
            .with_entities(
                db.func.to_char(FMCInformation.created_at, 'MM').label('month'),
                PipeInformation.pipe_size_inches,
                db.func.count(PipeInformation.id)
            )
            .group_by('month', PipeInformation.pipe_size_inches)
            .order_by('month')
            .all()
        )
        pipe_sizes = sorted(set(row[1] for row in pipe_size_month_counts), key=float)
        pipe_size_month_data = {str(size): [0] * 12 for size in pipe_sizes}
        for month, size, count in pipe_size_month_counts:
            pipe_size_month_data[str(size)][int(month) - 1] = count

        logger.debug("Analytics prepared.")

        return render_template(
            'index.html',
            entries=entries,
            total_entries=total_entries,
            longhaul_count=longhaul_count,
            gpon_fmc_count=gpon_fmc_count,
            total_cable_used=total_cable_used,
            total_joints=total_joints,
            total_pipe_used=total_pipe_used,
            month_labels=month_labels,
            month_counts=month_data,
            year_labels=year_labels,
            year_counts=year_data,
            cable_capacity_labels=cable_capacity_labels,
            cable_capacity_counts=cable_capacity_data,
            cable_type_counts=cable_type_counts,
            pipe_size_counts=pipe_size_counts,
            cable_types=cable_types,
            cable_type_month_data=cable_type_month_data,
            pipe_sizes=[str(size) for size in pipe_sizes],
            pipe_size_month_data=pipe_size_month_data,
            noc_by_category=noc_by_category_dict,
            noc_category_year_labels=noc_category_year_labels,
            noc_category_year_cats=noc_category_year_cats,
            noc_category_year_data=noc_category_year_data,
            noc_category_month_cats=noc_category_month_cats,
            noc_category_month_data=noc_category_month_data,
            noc_by_domain=noc_by_domain_dict,
            noc_domain_year_labels=noc_domain_year_labels,
            noc_domain_year_doms=noc_domain_year_doms,
            noc_domain_year_data=noc_domain_year_data,
            noc_domain_month_doms=noc_domain_month_doms,
            noc_domain_month_data=noc_domain_month_data,
            user_role=user_role,
            user_categories=user_categories,
            user_domains=user_domains
        )
    except Exception as e:
        logger.error(f"Error in index route: {str(e)}")
        flash(f"Error: {str(e)}")
        return redirect(url_for('index'))

@app.route('/add', methods=['GET', 'POST'])
@login_required
def add():
    try:
        user_role = session.get('role')
        user_categories = session.get('category', [])
        user_domains = session.get('domain', [])
        username = session.get('username', 'unknown_user')

        logger.debug(f"Add route: User={username}, Role={user_role}, Categories={user_categories}, Domains={user_domains}")

        if request.method == 'POST':
            category = request.form.get('category')
            domain = request.form.get('domain')
            cable_cut_noc_id = request.form.get('cable_cut_noc_id') or None
            cable_used_meters = safe_float(request.form.get('cable_used_meters'), 'cable_used_meters') if request.form.get('cable_used_meters') else None
            cable_type = request.form.get('cable_type') or None
            cable_capacity = request.form.get('cable_capacity') or None
            no_of_joints = safe_int(request.form.get('no_of_joints'), 'no_of_joints') if request.form.get('no_of_joints') else None

            if user_role != 'master':
                if category not in user_categories or domain not in user_domains:
                    logger.error(f"Non-master user {username} attempted to add data for category={category}, domain={domain}")
                    raise ValueError("You can only add data for your assigned categories and domains")
            else:
                if category not in VALID_CATEGORIES:
                    logger.error(f"Invalid category selected by master user: {category}")
                    raise ValueError("Invalid category")
                if domain not in VALID_DOMAINS[category]:
                    logger.error(f"Invalid domain selected by master user: category={category}, domain={domain}")
                    raise ValueError("Please select a specific domain for the chosen category")

            # Validate joint types
            joint_types = request.form.getlist('joint_type[]')
            joint_types = [jt.strip() for jt in joint_types if jt.strip()]
            if no_of_joints is not None and len(joint_types) != no_of_joints:
                raise ValueError(f"Number of joint types ({len(joint_types)}) does not match number of joints ({no_of_joints})")

            fmc = FMCInformation(
                region='RTR',
                category=category,
                domain=domain,
                cable_cut_noc_id=cable_cut_noc_id,
                cable_used_meters=cable_used_meters,
                cable_type=cable_type,
                cable_capacity=cable_capacity,
                no_of_joints=no_of_joints,
                created_by=username,
                updated_by=username,
                updated_at=datetime.utcnow()
            )
            db.session.add(fmc)
            db.session.flush()

            # Joint Types
            for jt in joint_types:
                joint = JointType(
                    fmc_id=fmc.id,
                    joint_type=jt,
                    created_by=username,
                    updated_by=username,
                    updated_at=datetime.utcnow()
                )
                db.session.add(joint)

            # Pipe Information
            pipe_used_meters = safe_float(request.form.get('pipe_used_meters'), 'pipe_used_meters') if request.form.get('pipe_used_meters') else None
            pipe_size_inches = safe_float(request.form.get('pipe_size_inches'), 'pipe_size_inches') if request.form.get('pipe_size_inches') else None
            pipe_type = request.form.get('pipe_type') or None
            if any([pipe_used_meters, pipe_size_inches, pipe_type]):
                pipe = PipeInformation(
                    fmc_id=fmc.id,
                    pipe_used_meters=pipe_used_meters,
                    pipe_size_inches=pipe_size_inches,
                    pipe_type=pipe_type,
                    created_by=username,
                    updated_by=username,
                    updated_at=datetime.utcnow()
                )
                db.session.add(pipe)

            db.session.commit()
            flash('Data added successfully!', 'success')
            return redirect(url_for('view_fmc'))
        else:
            logger.debug("Rendering add.html for GET request")
            return render_template('add.html', user_role=user_role, user_categories=user_categories, user_domains=user_domains)
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error in add route: {str(e)}")
        flash(f"Error adding data: {str(e)}", 'error')
        return render_template('add.html', user_role=user_role, user_categories=user_categories, user_domains=user_domains)
    
@app.route('/view_fmc', methods=['GET'])
@login_required
def view_fmc():
    try:
        page = request.args.get('page', 1, type=int)
        search = request.args.get('search', '').strip()

        query = FMCInformation.query
        query = apply_data_filter(query)

        if search:
            query = query.filter(FMCInformation.cable_cut_noc_id.ilike(f'%{search}%'))

        per_page = 10
        pagination = query.order_by(FMCInformation.created_at.desc()).paginate(page=page, per_page=per_page, error_out=False)

        # Convert timestamps to PKT for each FMC entry
        pkt_tz = pytz.timezone('Asia/Karachi')
        for fmc in pagination.items:
            fmc.created_at_formatted = fmc.created_at.replace(tzinfo=pytz.UTC).astimezone(pkt_tz).strftime('%-m/%-d/%Y, %-I:%M:%S %p')
            fmc.updated_at_formatted = fmc.updated_at.replace(tzinfo=pytz.UTC).astimezone(pkt_tz).strftime('%-m/%-d/%Y, %-I:%M:%S %p')

        # ✅ Extract distinct years from created_at column
        years = (
            FMCInformation.query
            .with_entities(db.extract('year', FMCInformation.created_at).label('year'))
            .distinct()
            .order_by('year')
            .all()
        )
        years = [int(year.year) for year in years]

        return render_template(
            'view_fmc.html',
            fmcs=pagination,
            search=search,
            user_role=session.get('role'),
            years=years
        )
    except Exception as e:
        logger.error(f"Error in view_fmc route: {str(e)}")
        flash(f"Error: {str(e)}")
        return redirect(url_for('index'))

@app.route('/edit/<int:id>', methods=['GET', 'POST'])
@master_required
def edit(id):
    try:
        user_categories = session.get('category', [])
        username = session.get('username', 'unknown_user')
        logger.debug(f"Edit route: User={username}, Editing ID={id}, User Categories={user_categories}")

        fmc = apply_data_filter(FMCInformation.query).filter_by(id=id).first_or_404()

        if request.method == 'POST':
            category = request.form.get('category')
            domain = request.form.get('domain')

            if category not in VALID_CATEGORIES:
                logger.error(f"Invalid category in edit: {category}")
                raise ValueError("Invalid category")
            if domain not in VALID_DOMAINS[category]:
                logger.error(f"Invalid domain in edit: category={category}, domain={domain}")
                raise ValueError("Please select a specific domain for the chosen category")

            fmc.category = category
            fmc.domain = domain
            fmc.cable_cut_noc_id = request.form.get('cable_cut_noc_id') or None
            fmc.cable_used_meters = safe_float(request.form.get('cable_used_meters'), 'cable_used_meters') if request.form.get('cable_used_meters') else None
            fmc.cable_type = request.form.get('cable_type') or None
            fmc.cable_capacity = request.form.get('cable_capacity') or None
            fmc.no_of_joints = safe_int(request.form.get('no_of_joints'), 'no_of_joints') if request.form.get('no_of_joints') else None
            fmc.updated_by = username
            fmc.updated_at = db.func.now()

            # Validate joint types
            joint_types = request.form.getlist('joint_type[]')
            joint_types = [jt.strip() for jt in joint_types if jt.strip()]
            if fmc.no_of_joints is not None and len(joint_types) != fmc.no_of_joints:
                raise ValueError(f"Number of joint types ({len(joint_types)}) does not match number of joints ({fmc.no_of_joints})")

            # Update Joint Types
            db.session.query(JointType).filter_by(fmc_id=fmc.id).delete()
            for jt in joint_types:
                joint = JointType(
                    fmc_id=fmc.id,
                    joint_type=jt,
                    created_by=fmc.updated_by,
                    updated_by=fmc.updated_by,
                    updated_at=db.func.now()
                )
                db.session.add(joint)

            # Update Pipe Information
            db.session.query(PipeInformation).filter_by(fmc_id=fmc.id).delete()
            pipe_used_meters = safe_float(request.form.get('pipe_used_meters'), 'pipe_used_meters') if request.form.get('pipe_used_meters') else None
            pipe_size_inches = safe_float(request.form.get('pipe_size_inches'), 'pipe_size_inches') if request.form.get('pipe_size_inches') else None
            pipe_type = request.form.get('pipe_type') or None
            if any([pipe_used_meters, pipe_size_inches, pipe_type]):
                pipe = PipeInformation(
                    fmc_id=fmc.id,
                    pipe_used_meters=pipe_used_meters,
                    pipe_size_inches=pipe_size_inches,
                    pipe_type=pipe_type,
                    created_by=fmc.updated_by,
                    updated_by=fmc.updated_by,
                    updated_at=db.func.now()
                )
                db.session.add(pipe)

            db.session.commit()
            flash('Data updated successfully!', 'success')
            return redirect(url_for('view_fmc'))
        else:
            joint_types = [jt.joint_type for jt in fmc.joint_types]
            pipe_info = fmc.pipe_info[0] if fmc.pipe_info else None
            return render_template('edit.html', fmc=fmc, joint_types=joint_types, pipe_info=pipe_info, user_categories=user_categories)
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error in edit route: {str(e)}")
        flash(f"Error editing data: {str(e)}", 'error')
        return render_template('edit.html', fmc=fmc, joint_types=joint_types, pipe_info=pipe_info, user_categories=user_categories)
    
@app.route('/delete/<int:id>', methods=['POST'])
@master_required
def delete(id):
    try:
        user_categories = session.get('category', [])
        username = session.get('username', 'unknown_user')
        logger.debug(f"Delete route: User={username}, Deleting ID={id}, User Categories={user_categories}")

        fmc = apply_data_filter(FMCInformation.query).filter_by(id=id).first_or_404()

        db.session.delete(fmc)
        db.session.commit()
        flash('Data deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error in delete route: {str(e)}")
        flash(f"Error deleting data: {str(e)}", 'error')
    return redirect(url_for('view_fmc'))

@app.route('/api/fmc/<int:id>', methods=['GET'])
@login_required
def get_fmc_details(id):
    try:
        fmc = apply_data_filter(FMCInformation.query).filter_by(id=id).first_or_404()
        pkt_tz = pytz.timezone('Asia/Karachi')

        created_at_formatted = fmc.created_at.replace(tzinfo=pytz.UTC).astimezone(pkt_tz).strftime('%d-%m-%Y, %I:%M %p') if fmc.created_at else "-"
        updated_at_formatted = fmc.updated_at.replace(tzinfo=pytz.UTC).astimezone(pkt_tz).strftime('%d-%m-%Y, %I:%M %p') if fmc.updated_at else "-"

        return jsonify({
            'id': fmc.id,
            'category': fmc.category,
            'domain': fmc.domain,
            'cable_cut_noc_id': fmc.cable_cut_noc_id,
            'cable_used_meters': fmc.cable_used_meters,
            'cable_type': fmc.cable_type,
            'cable_capacity': fmc.cable_capacity,
            'no_of_joints': fmc.no_of_joints,
            'created_by': fmc.created_by,
            'updated_by': fmc.updated_by,
            'created_at_formatted': created_at_formatted,
            'updated_at_formatted': updated_at_formatted,
            'joint_types': [{'id': jt.id, 'joint_type': jt.joint_type} for jt in fmc.joint_types],
            'pipe_info': [{
                'id': pi.id,
                'pipe_used_meters': pi.pipe_used_meters,
                'pipe_size_inches': pi.pipe_size_inches,
                'pipe_type': pi.pipe_type
            } for pi in fmc.pipe_info]
        })
    except Exception as e:
        logger.error(f"Error in get_fmc_details: {str(e)}")
        return jsonify({'error': str(e)}), 500

# ... (other imports and configurations remain unchanged)

@app.route('/export_fmc', methods=['GET'])
@login_required
def export_fmc():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.chart import BarChart, Reference, Series
        from openpyxl.chart.axis import TextAxis, ChartLines

        search = request.args.get('search', '').strip()
        years = request.args.getlist('year')  # Get list of selected years
        months = request.args.getlist('month')  # Get list of selected months
        query = FMCInformation.query
        query = apply_data_filter(query)

        if search:
            query = query.filter(FMCInformation.cable_cut_noc_id.ilike(f'%{search}%'))

        if years and 'All' not in years:
            years = [int(year) for year in years]  # Convert to integers for query
            query = query.filter(db.extract('year', FMCInformation.created_at).in_(years))
        if months and 'All' not in months:
            months = [month.zfill(2) for month in months]  # Ensure two-digit format
            query = query.filter(db.func.to_char(FMCInformation.created_at, 'MM').in_(months))

        data = query.order_by(FMCInformation.created_at.desc()).all()
        pkt_tz = pytz.timezone('Asia/Karachi')

        # Prepare Data sheet records
        data_records = []
        for fmc in data:
            created_at_pkt = fmc.created_at.replace(tzinfo=pytz.UTC).astimezone(pkt_tz)
            updated_at_pkt = fmc.updated_at.replace(tzinfo=pytz.UTC).astimezone(pkt_tz)
            joint_types_str = ', '.join(jt.joint_type for jt in fmc.joint_types) if fmc.joint_types else '-'
            pipe_info = fmc.pipe_info[0] if fmc.pipe_info else None
            pipe_used_meters = f'{pipe_info.pipe_used_meters:.2f}' if pipe_info and pipe_info.pipe_used_meters else '-'
            pipe_size_inches = f'{pipe_info.pipe_size_inches:.2f}' if pipe_info and pipe_info.pipe_size_inches else '-'
            pipe_type = pipe_info.pipe_type if pipe_info and pipe_info.pipe_type else '-'

            data_records.append({
                'ID': fmc.id,
                'Category': fmc.category,
                'Domain': fmc.domain,
                'NOC ID': fmc.cable_cut_noc_id or '-',
                'Cable Used (m)': f'{fmc.cable_used_meters:.2f}' if fmc.cable_used_meters else '-',
                'Cable Type': fmc.cable_type or '-',
                'Cable Capacity': fmc.cable_capacity or '-',
                'No. of Joints': fmc.no_of_joints if fmc.no_of_joints is not None else '-',
                'Joint Types': joint_types_str,
                'Pipe Used (m)': pipe_used_meters,
                'Pipe Size (in)': pipe_size_inches,
                'Pipe Type': pipe_type,
                'Created By': fmc.created_by,
                'Created At': created_at_pkt.strftime('%-m/%-d/%Y, %-I:%M:%S %p'),
                'Updated By': fmc.updated_by,
                'Updated At': updated_at_pkt.strftime('%-m/%-d/%Y, %-I:%M:%S %p')
            })

        # Prepare Summary sheet data
        summary_records = []
        chart_data = {}
        
        if data:
            # Total NOC IDs
            total_noc_ids = len(data)
            longhaul_count = sum(1 for fmc in data if fmc.category == 'Longhaul')
            gpon_fmc_count = sum(1 for fmc in data if fmc.category == 'GPON_FMC')

            # Total Cable Used
            total_cable_used = sum(fmc.cable_used_meters or 0 for fmc in data)

            # Total Joints
            total_joints = sum(fmc.no_of_joints or 0 for fmc in data)

            # Total Pipe Used
            pipe_query = apply_data_filter(db.session.query(PipeInformation).join(FMCInformation))
            if years and 'All' not in years:
                pipe_query = pipe_query.filter(db.extract('year', FMCInformation.created_at).in_(years))
            if months and 'All' not in months:
                pipe_query = pipe_query.filter(db.func.to_char(FMCInformation.created_at, 'MM').in_(months))
            total_pipe_used = pipe_query.with_entities(db.func.coalesce(db.func.sum(PipeInformation.pipe_used_meters), 0)).scalar()

            # Cable Type Counts
            cable_type_counts = Counter(fmc.cable_type for fmc in data if fmc.cable_type)
            cable_type_entries = sum(cable_type_counts.values())

            # Pipe Size Counts
            pipe_sizes = [pi.pipe_size_inches for pi in pipe_query.all() if pi.pipe_size_inches is not None]
            pipe_size_counts = Counter(pipe_sizes)
            pipe_size_entries = sum(pipe_size_counts.values())

            # NOC ID by Category
            noc_by_category = (
                query.with_entities(FMCInformation.category, db.func.count(FMCInformation.cable_cut_noc_id))
                .group_by(FMCInformation.category)
                .all()
            )
            chart_data['noc_by_category'] = {
                'labels': [row[0] for row in noc_by_category],
                'values': [row[1] for row in noc_by_category]
            }

            # NOC ID by Domain
            noc_by_domain = (
                query.with_entities(FMCInformation.domain, db.func.count(FMCInformation.cable_cut_noc_id))
                .group_by(FMCInformation.domain)
                .all()
            )
            chart_data['noc_by_domain'] = {
                'labels': [row[0] for row in noc_by_domain],
                'values': [row[1] for row in noc_by_domain]
            }

            # NOC ID Count by Month (2025)
            month_counts = (
                query.filter(db.extract('year', FMCInformation.created_at) == 2025)
                .with_entities(
                    db.func.to_char(FMCInformation.created_at, 'MM').label('month'),
                    db.func.count(FMCInformation.cable_cut_noc_id)
                )
                .group_by('month')
                .order_by('month')
                .all()
            )
            month_labels = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
            month_data = [0] * 12
            for month, count in month_counts:
                month_data[int(month) - 1] = count
            chart_data['noc_by_month'] = {
                'labels': month_labels,
                'values': month_data
            }

            # NOC ID Count by Year
            year_counts = (
                query.with_entities(
                    db.extract('year', FMCInformation.created_at).label('year'),
                    db.func.count(FMCInformation.cable_cut_noc_id)
                )
                .group_by('year')
                .all()
            )
            year_labels = sorted(set(str(int(year)) for year, _ in year_counts))
            year_data = [0] * len(year_labels)
            for year, count in year_counts:
                year_data[year_labels.index(str(int(year)))] = count
            chart_data['noc_by_year'] = {
                'labels': year_labels,
                'values': year_data
            }

            # Cable Capacity Distribution
            cable_capacity_rows = (
                query.filter(FMCInformation.cable_capacity.isnot(None))
                .with_entities(FMCInformation.cable_capacity, db.func.coalesce(db.func.sum(FMCInformation.cable_used_meters), 0))
                .group_by(FMCInformation.cable_capacity)
                .all()
            )
            chart_data['cable_capacity'] = {
                'labels': [row[0] for row in cable_capacity_rows],
                'values': [float(row[1]) for row in cable_capacity_rows]
            }

            # Build summary rows
            summary_records.append({'Metric': 'Total NOC IDs', 'Value': total_noc_ids, 'Details': ''})
            summary_records.append({'Metric': 'Longhaul', 'Value': longhaul_count, 'Details': ''})
            summary_records.append({'Metric': 'GPON_FMC', 'Value': gpon_fmc_count, 'Details': ''})
            summary_records.append({'Metric': 'Total Cable Used', 'Value': f'{total_cable_used:.2f} m', 'Details': 'Total Meters of Cable Deployed'})
            summary_records.append({'Metric': 'Total Joints', 'Value': total_joints, 'Details': 'Number of Joints Used'})
            summary_records.append({'Metric': 'Total Pipe Used', 'Value': f'{total_pipe_used:.2f} m', 'Details': 'Total Meters of Pipe Deployed'})
            summary_records.append({'Metric': 'Entries by Cable Type', 'Value': cable_type_entries, 'Details': ''})
            
            for cable_type, count in cable_type_counts.items():
                summary_records.append({'Metric': cable_type, 'Value': count, 'Details': ''})
            
            summary_records.append({'Metric': 'Entries by Pipe Size', 'Value': pipe_size_entries, 'Details': ''})
            for pipe_size, count in pipe_size_counts.items():
                summary_records.append({'Metric': f'{pipe_size:.2f} in', 'Value': count, 'Details': ''})
        else:
            # Initialize empty values when no data
            total_noc_ids = 0
            longhaul_count = 0
            gpon_fmc_count = 0
            total_cable_used = 0
            total_joints = 0
            total_pipe_used = 0
            cable_type_counts = Counter()
            pipe_size_counts = Counter()
            cable_type_entries = 0
            pipe_size_entries = 0

        # Create Excel workbook
        wb = Workbook()
        summary_ws = wb.active
        summary_ws.title = 'Summary'
        data_ws = wb.create_sheet('FMC Data')

        # Styling definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')

        # Data sheet: Write and style headers
        data_df = pd.DataFrame(data_records)
        if not data_df.empty:
            headers = list(data_records[0].keys())
            for col_num, header in enumerate(headers, 1):
                cell = data_ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border

            # Write data and apply alignment
            for row_num, row_data in enumerate(data_records, 2):
                for col_num, value in enumerate(row_data.values(), 1):
                    cell = data_ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.alignment = center_align
                    cell.border = border

            # Adjust column widths
            for col in data_ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                data_ws.column_dimensions[column].width = adjusted_width

        # Summary sheet: Create box-like layout
        summary_df = pd.DataFrame(summary_records)
        current_row = 2
        box_width = 4
        box_height = 4
        boxes_per_row = 3
        box_spacing = 2

        # Define summary boxes data
        boxes = [
            {
                'title': 'Total NOC IDs', 
                'value': total_noc_ids, 
                'details': '', 
                'subitems': [
                    {'label': 'Longhaul', 'value': longhaul_count},
                    {'label': 'GPON_FMC', 'value': gpon_fmc_count}
                ]
            },
            {
                'title': 'Total Cable Used', 
                'value': f'{total_cable_used:.2f} m', 
                'details': 'Total Meters of Cable Deployed'
            },
            {
                'title': 'Total Joints', 
                'value': total_joints, 
                'details': 'Number of Joints Used'
            },
            {
                'title': 'Total Pipe Used', 
                'value': f'{total_pipe_used:.2f} m', 
                'details': 'Total Meters of Pipe Deployed'
            },
            {
                'title': 'Entries by Cable Type', 
                'value': cable_type_entries, 
                'subitems': [
                    {'label': k, 'value': v} for k, v in cable_type_counts.items()
                ]
            },
            {
                'title': 'Entries by Pipe Size', 
                'value': pipe_size_entries, 
                'subitems': [
                    {'label': f'{k:.2f} in', 'value': v} for k, v in pipe_size_counts.items()
                ]
            }
        ]

        # Write summary boxes
        for idx, box in enumerate(boxes):
            col_start = 2 + (idx % boxes_per_row) * (box_width + box_spacing)
            row_start = current_row + (idx // boxes_per_row) * (box_height + box_spacing)

            # Title
            title_cell = summary_ws.cell(row=row_start, column=col_start)
            title_cell.value = box['title']
            title_cell.font = Font(bold=True, size=12)
            title_cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            title_cell.alignment = center_align
            summary_ws.merge_cells(start_row=row_start, start_column=col_start, end_row=row_start, end_column=col_start + box_width - 1)
            for col in range(col_start, col_start + box_width):
                summary_ws.cell(row=row_start, column=col).border = border

            # Value
            value_cell = summary_ws.cell(row=row_start + 1, column=col_start)
            value_cell.value = box['value']
            value_cell.font = Font(bold=True, size=14, color="00008B")
            value_cell.alignment = center_align
            summary_ws.merge_cells(start_row=row_start + 1, start_column=col_start, end_row=row_start + 1, end_column=col_start + box_width - 1)
            for col in range(col_start, col_start + box_width):
                summary_ws.cell(row=row_start + 1, column=col).border = border

            # Details or subitems
            if 'subitems' in box:
                for i, subitem in enumerate(box['subitems']):
                    label_cell = summary_ws.cell(row=row_start + 2 + i, column=col_start)
                    label_cell.value = f"{subitem['label']}: {subitem['value']}"
                    label_cell.alignment = left_align
                    label_cell.border = border
                    summary_ws.merge_cells(start_row=row_start + 2 + i, start_column=col_start, end_row=row_start + 2 + i, end_column=col_start + box_width - 1)
                    for col in range(col_start, col_start + box_width):
                        summary_ws.cell(row=row_start + 2 + i, column=col).border = border
            else:
                details_cell = summary_ws.cell(row=row_start + 2, column=col_start)
                details_cell.value = box.get('details', '')
                details_cell.alignment = left_align
                details_cell.border = border
                summary_ws.merge_cells(start_row=row_start + 2, start_column=col_start, end_row=row_start + 2, end_column=col_start + box_width - 1)
                for col in range(col_start, col_start + box_width):
                    summary_ws.cell(row=row_start + 2, column=col).border = border

            # Apply borders to entire box
            for row in range(row_start, row_start + box_height):
                for col in range(col_start, col_start + box_width):
                    cell = summary_ws.cell(row=row, column=col)
                    cell.border = border

        # Update current_row for next set of boxes
        current_row = 2 + ((len(boxes) - 1) // boxes_per_row + 1) * (box_height + box_spacing)

        # Add charts to Summary sheet only if data exists
        if data and chart_data:
            # Chart positions to prevent overlap
            chart_width = 15
            chart_height = 10

            # Chart 1: NOC ID by Category
            if 'noc_by_category' in chart_data:
                chart1 = BarChart()
                chart1.title = "NOC ID by Category"
                chart1.style = 10
                chart1.height = chart_height
                chart1.width = chart_width
                chart1_data = [['Category', 'Count']] + list(zip(chart_data['noc_by_category']['labels'], chart_data['noc_by_category']['values']))
                for row_num, row_data in enumerate(chart1_data, start=1):
                    for col_num, value in enumerate(row_data, 1):
                        summary_ws.cell(row=row_num + 200, column=col_num).value = value
                data_ref = Reference(summary_ws, min_col=2, min_row=201, max_row=200 + len(chart1_data), max_col=2)
                cats_ref = Reference(summary_ws, min_col=1, min_row=202, max_row=200 + len(chart1_data))
                chart1.add_data(data_ref, titles_from_data=True)
                chart1.set_categories(cats_ref)
                chart1.dataLabels = DataLabelList()
                chart1.dataLabels.showVal = True
                chart1.dataLabels.showCatName = False
                chart1.dataLabels.showSerName = False
                chart1.dataLabels.showLegendKey = False
                chart1.y_axis.title = "Count"
                chart1.x_axis.title = "Category"
                chart1.y_axis.majorGridlines = ChartLines()
                summary_ws.add_chart(chart1, "A15")

            # Chart 2: NOC ID by Domain
            if 'noc_by_domain' in chart_data:
                chart2 = BarChart()
                chart2.title = "NOC ID by Domain"
                chart2.style = 10
                chart2.height = chart_height
                chart2.width = chart_width
                chart2_data = [['Domain', 'Count']] + list(zip(chart_data['noc_by_domain']['labels'], chart_data['noc_by_domain']['values']))
                for row_num, row_data in enumerate(chart2_data, start=1):
                    for col_num, value in enumerate(row_data, 4):
                        summary_ws.cell(row=row_num + 200, column=col_num).value = value
                data_ref2 = Reference(summary_ws, min_col=5, min_row=201, max_row=200 + len(chart2_data), max_col=5)
                cats_ref2 = Reference(summary_ws, min_col=4, min_row=202, max_row=200 + len(chart2_data))
                chart2.add_data(data_ref2, titles_from_data=True)
                chart2.set_categories(cats_ref2)
                chart2.dataLabels = DataLabelList()
                chart2.dataLabels.showVal = True
                chart2.dataLabels.showCatName = False
                chart2.dataLabels.showSerName = False
                chart2.dataLabels.showLegendKey = False
                chart2.y_axis.title = "Count"
                chart2.x_axis.title = "Domain"
                chart2.y_axis.majorGridlines = ChartLines()
                summary_ws.add_chart(chart2, "J15")

            # Chart 3: NOC ID Count by Month (2025)
            if 'noc_by_month' in chart_data:
                chart3 = BarChart()
                chart3.title = "NOC ID Count by Month (2025)"
                chart3.style = 10
                chart3.height = chart_height
                chart3.width = chart_width
                chart3_data = [['Month', 'Count']] + list(zip(chart_data['noc_by_month']['labels'], chart_data['noc_by_month']['values']))
                for row_num, row_data in enumerate(chart3_data, start=1):
                    for col_num, value in enumerate(row_data, 7):
                        summary_ws.cell(row=row_num + 200, column=col_num).value = value
                data_ref3 = Reference(summary_ws, min_col=8, min_row=201, max_row=200 + len(chart3_data), max_col=8)
                cats_ref3 = Reference(summary_ws, min_col=7, min_row=202, max_row=200 + len(chart3_data))
                chart3.add_data(data_ref3, titles_from_data=True)
                chart3.set_categories(cats_ref3)
                chart3.dataLabels = DataLabelList()
                chart3.dataLabels.showVal = True
                chart3.dataLabels.showCatName = False
                chart3.dataLabels.showSerName = False
                chart3.dataLabels.showLegendKey = False
                chart3.y_axis.title = "Count"
                chart3.x_axis.title = "Month"
                chart3.y_axis.majorGridlines = ChartLines()
                summary_ws.add_chart(chart3, "A36")

            # Chart 4: NOC ID Count by Year
            if 'noc_by_year' in chart_data:
                chart4 = BarChart()
                chart4.title = "NOC ID Count by Year"
                chart4.style = 10
                chart4.height = chart_height
                chart4.width = chart_width
                chart4_data = [['Year', 'Count']] + list(zip(chart_data['noc_by_year']['labels'], chart_data['noc_by_year']['values']))
                for row_num, row_data in enumerate(chart4_data, start=1):
                    for col_num, value in enumerate(row_data, 10):
                        summary_ws.cell(row=row_num + 200, column=col_num).value = value
                data_ref4 = Reference(summary_ws, min_col=11, min_row=201, max_row=200 + len(chart4_data), max_col=11)
                cats_ref4 = Reference(summary_ws, min_col=10, min_row=202, max_row=200 + len(chart4_data))
                chart4.add_data(data_ref4, titles_from_data=True)
                chart4.set_categories(cats_ref4)
                chart4.dataLabels = DataLabelList()
                chart4.dataLabels.showVal = True
                chart4.dataLabels.showCatName = False
                chart4.dataLabels.showSerName = False
                chart4.dataLabels.showLegendKey = False
                chart4.y_axis.title = "Count"
                chart4.x_axis.title = "Year"
                chart4.y_axis.majorGridlines = ChartLines()
                summary_ws.add_chart(chart4, "J36")

            # Chart 5: Cable Capacity Distribution
            if 'cable_capacity' in chart_data:
                chart5 = BarChart()
                chart5.title = "Cable Capacity Distribution"
                chart5.style = 10
                chart5.height = chart_height
                chart5.width = chart_width
                chart5_data = [['Cable Capacity', 'Meters']] + list(zip(chart_data['cable_capacity']['labels'], chart_data['cable_capacity']['values']))
                for row_num, row_data in enumerate(chart5_data, start=1):
                    for col_num, value in enumerate(row_data, 13):
                        summary_ws.cell(row=row_num + 200, column=col_num).value = value
                data_ref5 = Reference(summary_ws, min_col=14, min_row=201, max_row=200 + len(chart5_data), max_col=14)
                cats_ref5 = Reference(summary_ws, min_col=13, min_row=202, max_row=200 + len(chart5_data))
                chart5.add_data(data_ref5, titles_from_data=True)
                chart5.set_categories(cats_ref5)
                chart5.dataLabels = DataLabelList()
                chart5.dataLabels.showVal = True
                chart5.dataLabels.showCatName = False
                chart5.dataLabels.showSerName = False
                chart5.dataLabels.showLegendKey = False
                chart5.y_axis.title = "Meters"
                chart5.x_axis.title = "Cable Capacity"
                chart5.y_axis.majorGridlines = ChartLines()
                summary_ws.add_chart(chart5, "H57")

        # Save workbook
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response
        filename_years = 'All' if not years or 'All' in years else '-'.join(str(year) for year in years)
        filename_months = 'All' if not months or 'All' in months else '-'.join(months)
        response = make_response(send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'fmc_data_{filename_years}_{filename_months}.xlsx'
        ))

        # Preserve auth_token
        auth_token = request.cookies.get('auth_token')
        if auth_token:
            response.set_cookie(
                'auth_token',
                auth_token,
                httponly=True,
                secure=True,
                samesite='Lax',
                max_age=int(timedelta(hours=24).total_seconds())
            )

        logger.info(f"Export successful for user {session.get('username')}, Years={filename_years}, Months={filename_months}")
        return response
        
    except Exception as e:
        logger.error(f"Error in export_fmc: {str(e)}")
        flash(f"Error exporting data: {str(e)}")
        return redirect(url_for('view_fmc'))

# ... (rest of the app.py remains unchanged)

# Add before_request hook
@app.before_request
def check_api_auth():
    if request.path.startswith('/api') and request.path not in ['/api/user_info', '/api/login']:
        if 'user_id' not in session:
            logger.error(f"Unauthorized API access: {request.path}, Headers={request.headers}")
            return jsonify({"error": "Unauthorized. Please login."}), 401

@app.route('/api/login', methods=['POST'])
@csrf.exempt
def api_login():
    logger.debug(f"Raw request: Method={request.method}, Headers={dict(request.headers)}, Form={request.form}, Data={request.get_data(as_text=True)}")

    content_type = request.headers.get('Content-Type', '').lower()
    if not content_type.startswith('application/x-www-form-urlencoded'):
        logger.error(f"Invalid Content-Type: {content_type or 'missing'}")
        return jsonify({"error": "Content-Type must be application/x-www-form-urlencoded"}), 415

    try:
        username = request.form.get('username', '').lower().strip()
        password = request.form.get('password', '').strip()

        if not username or not password:
            logger.error("Missing username or password")
            return jsonify({"error": "Username and password are required"}), 400

        user_data = supabase_service.table('users_info').select('user_id', 'region', 'category', 'domain', 'role').eq('username', username).execute()
        if not user_data.data:
            logger.error(f"Invalid username: {username}")
            return jsonify({"error": "Invalid username"}), 401

        user_id = user_data.data[0]['user_id']
        region = user_data.data[0]['region']
        categories = user_data.data[0]['category']
        domains = user_data.data[0]['domain']
        role = user_data.data[0]['role']

        auth_user = supabase_service.auth.admin.get_user_by_id(user_id)
        if not auth_user.user:
            logger.error(f"User not found in authentication system: {username}")
            return jsonify({"error": "User not found in authentication system"}), 401

        email = auth_user.user.email
        email_confirmed = bool(auth_user.user.email_confirmed_at)

        if not email_confirmed:
            supabase.auth.resend({"type": "signup", "email": email})
            logger.info(f"Email not verified for {username}, resending verification email")
            return jsonify({"error": "Please verify your email. A new verification email has been sent."}), 403

        response = supabase.auth.sign_in_with_password({"email": email, "password": password})
        if response.user and response.user.id == user_id:
            session['region'] = region
            session['username'] = username
            session['user_id'] = user_id
            session['category'] = categories
            session['domain'] = domains
            session['role'] = role
            resp = make_response(jsonify({
                "message": "Login successful",
                "auth_token": response.session.access_token,
                "user": {
                    "username": username,
                    "region": region,
                    "category": categories,
                    "domain": domains,
                    "role": role
                }
            }))
            resp.set_cookie(
                "auth_token",
                response.session.access_token,
                httponly=True,
                secure=True,
                samesite="Lax",
                max_age=int(timedelta(hours=24).total_seconds())
            )
            logger.info(f"API login successful for {username}")
            return resp
        else:
            logger.error(f"Invalid credentials for {username}")
            return jsonify({"error": "Invalid credentials"}), 401
    except Exception as e:
        logger.error(f"Error in api_login: {str(e)}")
        return jsonify({"error": f"Login failed: {str(e)}"}), 500

# Add /api/user_info
@app.route('/api/user_info', methods=['GET'])
@csrf.exempt
def get_user_info():
    try:
        username = request.args.get('username', '').lower().strip()
        if not username:
            logger.error("Username is required")
            return jsonify({'error': 'Username is required'}), 400

        user_data = supabase_service.table('users_info').select('email').eq('username', username).execute()
        if not user_data.data:
            logger.error(f"Username not found: {username}")
            return jsonify({'error': 'Username not found'}), 404

        return jsonify({'email': user_data.data[0]['email']}), 200
    except Exception as e:
        logger.error(f"Error in get_user_info: {str(e)}")
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True)